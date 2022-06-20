import openpyxl

#criar uma planilha(book)

book = openpyxl.Workbook()

# visializar paginas
print(book.sheetnames)

# criar paginas

book.create_sheet("frutas")

#sselecionar a pagina

frutas_page = book["frutas"]
frutas_page.append(["Nome", "Qtd", "Valor R$"])
frutas_page.append(["banana", "5", "R$3,90"])
frutas_page.append(["fruta2", "3", "R$65,90"])
frutas_page.append(["fruta3", "5", "R$7,10"])
frutas_page.append(["fruta4", "10", "R$5,80"])

#salvar palnilha

book.save("Planilha de compras.xlsx")